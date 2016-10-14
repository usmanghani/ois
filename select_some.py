import os
import openpyxl
import re
import csv
import cStringIO as StringIO
import sys


IS_IB = os.environ.get('INSTABASE_URI', None) is not None
RELATIVE_PATH = '/ughani/my-repo/fs/Instabase Drive/files/' if IS_IB else ''
CHUNK_SIZE = 10485760 if IS_IB else -1  # read whole file when not in ib

# open_method = ib.open if IS_IB else open
open_method = open


def chunked_read(filehandle):
    if not IS_IB: return filehandle.read()
    data = []
    while filehandle.tell() != -1:
        data.append(filehandle.read(CHUNK_SIZE))  # reads only 10M at a time.
    return ''.join(data)


read_chunked = chunked_read


delim_regex = r'''(?P<CN>\d+),HPD(?P<DB>.*?)\s*Incident\s*#.*?(?P<IN>\d+)'''
compiled = re.compile(delim_regex)

NUM_CASES_TO_SELECT = int(sys.argv[1]) if len(sys.argv) > 1 and sys.argv[1] and str.isdigit(sys.argv[1]) else 100

with open_method('brad_houston_files/raw_narratives_20160210.txt') as houston_file:
    contents = houston_file.read()

last_match = None
counter = 0
pos = 0
MAX_NARRATIVES = NUM_CASES_TO_SELECT or 100

for current_match in compiled.finditer(contents):
    if not last_match:
        last_match = current_match
        pos = last_match.end()
        continue

    counter += 1
    if counter > MAX_NARRATIVES: break

    pos = current_match.start()
    last_match = current_match

with open_method('brad_houston_files/raw_narratives_20160210-extracted.txt', 'w') as narrative_output:
    narrative_output.write(contents[0:pos])


MAX_STRUCTURED_RECORDS = NUM_CASES_TO_SELECT or 100


with open_method(RELATIVE_PATH + 'brad_houston_files/merged_raw_suspect_officer.xlsx') as structured_file:
    contents = StringIO.StringIO(chunked_read(structured_file))

    houston_input_vars_wb = openpyxl.load_workbook(contents, data_only=True, read_only=True)
    first_ws = houston_input_vars_wb.worksheets[0]
    first_row = next(first_ws.rows)
    names = [cell.value for cell in first_row]

    with open_method('brad_houston_files/merged_raw_suspect_officer-extracted.csv', 'w') as output_structured:
        writer = csv.writer(output_structured)
        writer.writerow(names)
        for row in first_ws.get_squared_range(1, 2, first_ws.max_column, min(first_ws.max_row, MAX_STRUCTURED_RECORDS)):
            writer.writerow([row[i].value for i in xrange(0, len(row))])

