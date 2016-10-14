import os
import openpyxl
import unicodecsv
import cStringIO as StringIO
import sys
import argparse
import re
import time
import xlrd


class ExcelReader(object):
    def __init__(self, file_path, skipheader=True):
        self.skipheader = skipheader

    def get_all_rows(self, ws):
        pass

    def get_max_row(self, ws):
        pass

    def get_worksheets(self):
        pass

    def get_worksheet_name(self, ws):
        pass

    def get_squared_range(self, ws, min_col, min_row, max_col, max_row):
        pass

    def get_max_col(self, ws):
        pass

class XlsReader(ExcelReader):
    def __init__(self, file_path, skipheader=True):
        super(XlsReader, self).__init__(file_path, skipheader)
        self.wb = xlrd.open_workbook(file_path)

    def get_worksheets(self):
        return self.wb.sheets()

    def get_max_row(self, ws):
        return ws.nrows

    def get_max_col(self, ws):
        return ws.ncols

    def get_all_rows(self, ws):
        return ws.get_rows()

    def get_worksheet_name(self, ws):
        return ws.name

    def get_squared_range(self, ws, min_col, min_row, max_col, max_row):
        """Returns a 2D array of cells

        :param min_col: smallest column index (1-based index)
        :type min_col: int

        :param min_row: smallest row index (1-based index)
        :type min_row: int

        :param max_col: largest column index (1-based index)
        :type max_col: int

        :param max_row: smallest row index (1-based index)
        :type max_row: int

        :rtype: generator
        """
        # Column name cache is very important in large files.
        for row in range(min_row - 1, max_row):
            yield tuple(ws.cell(rowx=row, colx=column)
                        for column in range(min_col - 1, max_col))


class XlsxReader(ExcelReader):
    def __init__(self, file_path, skipheader=True):
        super(XlsxReader, self).__init__(file_path, skipheader)
        with open(file_path) as excel_file:
            contents = StringIO.StringIO(excel_file.read())
            self.wb = openpyxl.load_workbook(contents, data_only=True, read_only=True)

    def get_worksheets(self):
        return self.wb.worksheets

    def get_max_row(self, ws):
        return ws.max_row

    def get_max_col(self, ws):
        return ws.max_column

    def get_all_rows(self, ws):
        return ws.rows

    def get_worksheet_name(self, ws):
        return ws.title

    def get_squared_range(self, ws, min_col, min_row, max_col, max_row):
        return ws.get_squared_range(min_col, min_row, max_col, max_row)


def hms_string(sec_elapsed):
    h = int(sec_elapsed / (60 * 60))
    m = int((sec_elapsed % (60 * 60)) / 60)
    s = sec_elapsed % 60.
    return "{}:{:>02}:{:>05.2f}".format(h, m, s)


def process_file(input_file, output_dir, skipheader, max_rows):
    start_time = time.time()

    print "Processing file %s" % input_file.encode('utf-8')

    input_filename, _ = os.path.splitext(os.path.basename(input_file))

    excel_reader = XlsxReader(input_file, skipheader) if input_file.endswith('xlsx') else XlsReader(input_file, skipheader)

    for worksheet in excel_reader.get_worksheets():
        total = (min(excel_reader.get_max_row(worksheet), max_rows) - 2 if skipheader else 1)

        print "About to process a total of %d rows (%s:%s)" % (total, input_filename, excel_reader.get_worksheet_name(worksheet))

        output_name = ''.join([input_filename, '_', excel_reader.get_worksheet_name(worksheet)])

        output_name = re.sub(r"\[\]\/\;,><&\*:%=\+@!#\^\(\)\|\?\^", '', output_name)

        if total <= 0:
            print "Sheet is empty. Skipping..."
            continue

        first_row = next(excel_reader.get_all_rows(worksheet))

        if skipheader:
            names = [cell.value for cell in first_row]

        output_path = os.path.join(output_dir, '.'.join([output_name, 'csv']))

        print "Writing output file %s" % output_path

        with open(output_path, 'w') as output_file:
            writer = unicodecsv.writer(output_file)

            if skipheader:
                writer.writerow(names)

            counter = 0

            rows_to_process = excel_reader.get_squared_range(
                    worksheet, 1, 2 if skipheader else 1,
                    excel_reader.get_max_col(worksheet), min(excel_reader.get_max_row(worksheet), max_rows))

            for row in rows_to_process:
                counter += 1

                if counter % 10000 == 0:
                    print "Processed %d / %d rows" % (counter, total)

                writer.writerow([row[i].value for i in xrange(0, len(row))])

    end_time = time.time()

    print "File %s took %s to process" % (input_file.encode('utf-8'), hms_string(end_time - start_time))


def process_dir(input, output_base, output_dir, skipheader, max_rows):
    start_time = time.time()

    print "Processing directory %s" % input.encode('utf-8')

    for root, dirnames, filenames in os.walk(input):

        files_to_process = [f for f in filenames if f.endswith('xlsx') or f.endswith('xls')]

        file_output_path = os.path.join(output_base + "/" + root)

        os.makedirs(file_output_path) if not os.path.exists(file_output_path) and files_to_process else None

        for filename in files_to_process:
            # print "Processing file %s and writing to %s" % (os.path.join(root, filename), file_output_path)
            process_file(os.path.join(root, filename), file_output_path, skipheader, max_rows)

        for dirname in dirnames:
            process_dir(os.path.join(root, dirname), output_base, output_dir, skipheader, max_rows)

    end_time = time.time()

    print "Directory %s took %s to process" % (input.encode('utf-8'), hms_string(end_time - start_time))

parser = argparse.ArgumentParser(prog='xlsx2csv', usage='%(prog)s [options]')

parser.add_argument('-i', '--input', dest='input', required=True)
parser.add_argument('-o', '--output', dest='output', required=True, help='output dir. if the input is a dir, '
                                                         'then output is considered a relative path and '
                                                         'each input folder has an output folder created '
                                                         'in-place for converted files.')
parser.add_argument('-ob', '--output-base', dest='output_base', required=False, help='Base directory to use for ouput when the '
                                                                                     'input directory is read-only.',
                    default='')

parser.add_argument('--dontskipheader', action='store_false', dest='skipheader', default=True)
parser.add_argument('--maxrows', type=int, dest='max_rows', default=99999)


args = parser.parse_args()
print args

if not os.path.exists(args.input):
    print "Input path doesn't exist"
    sys.exit(1)

if os.path.exists(args.output) and not os.path.isdir(args.output):
    print "Output path exists and is not a dir"
    sys.exit(1)

if args.output_base and os.path.exists(args.output_base) and not os.path.isdir(args.output_base):
    print "Base output path exists and is not a dir"
    sys.exit(1)

if args.output_base and not os.path.exists(args.output_base):
    os.makedirs(args.output_base)

if os.path.isdir(args.input):
    process_dir(args.input, args.output_base, args.output, args.skipheader, args.max_rows)
else:
    if not os.path.exists(args.output):
        os.makedirs(args.output)
    process_file(args.input, os.path.join(args.output_base, args.output), args.skipheader, args.max_rows)




